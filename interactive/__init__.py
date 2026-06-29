from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from utils import (
    delay,
    get_random_user_agent,
    get_xlsx_filepath,
    save_xlsx,
)
from worker import get_data_by_worker_id, get_xlsx_data, write_csv_by_id

from .etf import funds_etf
from .it import funds_investment
from .mutual_fund import funds_mf, get_total_funds_mf

#
# def get_funds_keywords(driver: WebDriver, funds: list[dict]) -> list[dict]:
#    data = []
#    wait = WebDriverWait(driver, timeout=5)
#    for fund in funds:
#        url = fund["url"]
#        get_with_backoff(driver, url)
#        print(f"current url: {url}")
#        keyword_xpath = '//div[@data-testid="allowedaccounts"]/p'
#        keyword_elm = find_element_or_none(wait, keyword_xpath)
#        keyword = None
#        if keyword_elm:
#            keyword = keyword_elm.text.strip()
#        f = dict(name=fund["name"], isin=fund["isin"], url=fund["url"], keyword=keyword)
#        data.append(f)
#        delay(2, 3)
#    return data
#


def create_fresh_browser(playwright_ctx):
    """Launches a clean browser with CI/CD optimizations."""
    browser = playwright_ctx.chromium.launch(
        headless=True, args=["--disable-dev-shm-usage", "--no-sandbox", "--disable-gpu"]
    )
    context = browser.new_context()
    page = context.new_page()
    return browser, context, page


def get_funds_keywords_playwright(
    playwright_ctx, funds: list[dict], sheet: str, batch_size: int = 50
) -> list[dict]:
    data = []

    # Initialize the first browser instance
    browser, context, page = create_fresh_browser(playwright_ctx)

    for index, fund in enumerate(funds):
        # --- Memory Leak Protection / Periodic Restart ---
        if index % batch_size == 0 and index != 0:
            print(
                f"\n--- Batch limit reached ({index}). Recycling browser to clear memory... ---"
            )
            context.close()
            browser.close()
            # Spin up a totally fresh browser process
            browser, context, page = create_fresh_browser(playwright_ctx)

        url = fund["url"]
        print(f"[{index + 1}/{len(funds)}] current url: {url}")

        keyword = None
        retries = 3

        # --- Network Backoff and Navigation Loop ---
        for attempt in range(retries):
            try:
                # Playwright's goto combines navigation and wait
                page.goto(url, timeout=30000, wait_until="domcontentloaded")

                # Dynamic wait targeting your specific React element via test-id
                selector = 'div[data-testid="allowedaccounts"] p'
                page.wait_for_selector(selector, timeout=5000)

                # Extract text
                keyword = page.locator(selector).first.inner_text().strip()
                break  # Success, break out of retry loop

            except Exception:
                if attempt == retries - 1:
                    print(f"Failed to scrape {url} after {retries} attempts.")
                else:
                    print(f"Timeout/Error. Retrying [{attempt + 2}] in 2 seconds...")
                    delay(1, 2)

        # Append data matching your dictionary schema
        f = dict(
            index=fund.get("index"),
            name=fund.get("name"),
            isin=fund.get("isin"),
            url=fund.get("url"),
            keyword=keyword,
            sheet=sheet,
        )
        data.append(f)

        # Mimic your original delay(2, 3) if you need pacing
        delay(1, 3)

    # Final clean up
    context.close()
    browser.close()

    return data


def write_spreadsheet(filepath: str, sheet: str, data: list[dict]) -> None:
    wb = load_workbook(filepath)
    ws = wb[sheet]
    i = 2
    for f in data:
        ws.cell(i, 1, f["name"])
        ws.cell(i, 2, f["isin"])
        cell = ws.cell(i, 3, f["url"])
        cell.style = "Hyperlink"
        cell.hyperlink = f["url"]
        ws.cell(i, 4, f["keyword"])
        i += 1
    wb.save(filepath)
    wb.close()


def get_urls(sheet: str) -> None:
    headers = {
        "accept": "application/json;charset=UTF-8",
        "accept-encoding": "gzip, deflate, br",
        "accept-language": "en-US, en;q=0.9,fr-FR;q=0.8,fr;q=0.7",
        "cache-control": "no-cache",
        "ii-consumer-type": "web.public",
        "origin": "https://www.ii.co.uk",
        "pragma": "no-cache",
    }
    ua = get_random_user_agent()
    headers.update(ua)
    xlsx_path = get_xlsx_filepath("interactive_investor.xlsx")
    data = []
    cols = ["name", "isin", "url"]
    for sheet in ["Investment", "ETF", "MF"]:
        if sheet == "Investment":
            data = funds_investment(headers)
            save_xlsx(xlsx_path, data, cols, sheet)

        if sheet == "ETF":
            data = funds_etf(headers)
            save_xlsx(xlsx_path, data, cols, sheet)

        if sheet == "MF":
            total_mf = get_total_funds_mf(headers)
            total_inc = total_mf["inc"]
            total_acc = total_mf["acc"]
            mf_data_inc = funds_mf(headers, total_inc, "inc")
            mf_data_acc = funds_mf(headers, total_acc, "acc")
            data = mf_data_acc + mf_data_inc
            save_xlsx(xlsx_path, data, cols, sheet)
    return


def interactive_runner(id_worker: int, max_worker: int, sheet: str):
    xlsx_path = get_xlsx_filepath("interactive_investor.xlsx")
    funds_data = get_xlsx_data(xlsx_path, sheet)
    worker_data = get_data_by_worker_id(id_worker, max_worker, funds_data)
    out_csv = f"ii_{id_worker}_{sheet}.csv"
    fields = ["index", "name", "isin", "url", "keyword", "sheet"]

    with sync_playwright() as pw_ctx:
        worker_data = get_funds_keywords_playwright(pw_ctx, worker_data, sheet)
        write_csv_by_id(out_csv, worker_data, fields)


def funds_dedup(funds: list[dict]) -> list[dict]:
    seen = set()
    dedup = [
        d
        for d in funds
        if tuple(d.items()) not in seen and not seen.add(tuple(d.items()))
    ]
    return dedup
