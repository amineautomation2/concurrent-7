import csv
import os
from collections.abc import Callable
from typing import Any

import openpyxl

from utils import save_xlsx


def get_xlsx_data(filename, sheet_name) -> list[dict]:
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]
    data = []
    row_start = 2
    for row in range(row_start, ws.max_row + 1):
        c1 = ws.cell(row, 1).value
        c2 = ws.cell(row, 2).value
        c3 = ws.cell(row, 3).value
        f = dict(index=row, name=c1, isin=c2, url=c3, sheet=sheet_name)
        data.append(f)
    wb.close()
    return data


def get_data_by_worker_id(
    id: int,
    max_worker: int,
    data: list,
) -> list[dict]:
    return data[id::max_worker]


def write_csv_by_id(
    filename: str,
    data: list[dict],
    fields: list[str],
) -> None:
    filepath = os.path.join("csv", filename)
    with open(filepath, "w+", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(data)
        print(f"file {filename} created.")


def write_csv(out: str, data: list[dict], fields: list[str]) -> None:
    with open(out, "w+", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(data)


def read_csv(filename: str) -> list[dict]:
    csv_data = []
    with open(filename, "r", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            csv_data.append(row)
    return csv_data


def process_data(
    fn: Callable[[Any], Any],
    data: list[dict],
) -> list[dict]:
    return fn(data)


def add_dummy_isin(data: list[dict]) -> list[dict]:
    new_data = []
    for f in data:
        f.update(dict(isin=f"isin_{f.get('index')}"))
        new_data.append(f)
    return new_data


def merge_csv_to_xlsx(xlsx_out: str, fields: list[str], sheet: str):
    # xlsx_in = os.path.join(os.getcwd(), "spreadsheet", "hl.xlsx")
    combined_data = []
    csv_dir = os.path.join(os.getcwd(), "csv")
    for filename in os.listdir(csv_dir):
        if filename.endswith(f"{sheet}.csv"):
            file_path = os.path.join(csv_dir, filename)
            data = read_csv(file_path)
            if data:
                combined_data.extend(data)
    # sorted_data = sorted(combined_data, key=lambda x: int(x["index"]))
    # write_csv(output_file, sorted_data, ["index", "name", "isin", "url"])
    for item in combined_data:
        item.pop("sheet", None)
    save_xlsx(xlsx_out, combined_data, fields, sheet)
    print(f"Successfully merged all files into {xlsx_out}")
