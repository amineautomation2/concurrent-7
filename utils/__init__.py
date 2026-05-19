from random import uniform
from time import sleep
import time
from curl_cffi import requests
import random
from pathlib import Path
from os.path import join, dirname, basename
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from ua_generator.options import Options as OptionsUA
from ua_generator.data.version import VersionRange
from datetime import datetime
import openpyxl
import json
import ua_generator


def get_random_user_agent() -> dict:
    options = OptionsUA()
    options.version_ranges = {
        'chrome': VersionRange(140, 144),  # Choose version between 125 and 129
    }
    ua = ua_generator.generate(
        browser='chrome', platform='windows', options=options)
    ua.headers.accept_ch(
        "Sec-CH-UA-Platform-Version, Sec-CH-UA-Full-Version-List")
    # return ua.headers.get()
    headers = ua.headers.get()
    return {k.title(): v for k, v in headers.items()}


def delay(min: float, max: float):
    sleep(uniform(min, max))


def setup_driver(headless: bool = False, pref: dict = {}) -> WebDriver:
    ua = get_random_user_agent()

    options = Options()
    if headless:
        options.add_argument("--headless=new")  # New headless mode

    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1920,1080")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("prefs", pref)
    options.add_argument("--disable-images")
    options.add_argument(f"--user-agent={ua.get("User-Agent")}")
    options.add_argument("--blink-settings=imagesEnabled=false")
    return webdriver.Chrome(options=options)


def find_element_or_none(wait: WebDriverWait, selector: str) -> WebElement | None:
    try:
        elm = wait.until(EC.presence_of_element_located((By.XPATH, selector)))
        return elm
    except:
        return None


def find_elements(wait: WebDriverWait, selector: str) -> list[WebElement] | None:
    try:
        children = wait.until(
            EC.visibility_of_any_elements_located((By.XPATH, selector))
        )
        return children
    except:
        return None


def clean_spreadsheet(filename: str) -> None:
    wb = openpyxl.load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                cell.value = None
                cell.hyperlink = None
    wb.save(filename)
    wb.close()


def get_current_quarter(date_obj) -> str:
    return f"Q{(date_obj.month - 1) // 3 + 1}"


def email_title() -> str:
    now = datetime.now()
    return f"[Automation] {get_current_quarter(now)} report for Halifax - Chelsea Financial - iWeb - Quilter - Standard Life - Willis Owen"


def get_xlsx_filepath(filename: str) -> str:
    project_root = Path(__file__).resolve().parent.parent
    return join(project_root, "spreadsheet", filename)


def write_json(filename: str, data: list[dict]) -> None:
    with open(filename, "w+") as f:
        json.dump(data, f, indent=4)


def read_json(filename: str):
    with open(filename, "r") as f:
        return json.load(f)


def fetch_with_backoff(url, headers=get_random_user_agent(), cookies=None, data=None, max_retries=5, base_delay=2):
    for attempt in range(max_retries):
        try:
            # Using curl_cffi to mimic a real browser (e.g., Chrome)
            response = requests.get(
                url, headers=headers, cookies=cookies, impersonate="chrome", timeout=10)

            # If successful, return the response
            if response.status_code == 200:
                return response

            # If we hit rate limits or server errors, we should retry
            if response.status_code in [429, 500, 502, 503, 504]:
                print(
                    f"Attempt {attempt + 1} failed with status {response.status_code}. Retrying...")
            else:
                # For 404 or 403, retrying usually won't help
                print(
                    f"Permanent error {response.status_code}. Skipping retries.")
                return response

        except Exception as e:
            print(f"Attempt {attempt + 1} raised an exception: {e}")

        jitter = random.uniform(0.5, 1.5)
        sleep_time = (base_delay * (2 ** attempt)) * jitter

        print(f"Sleeping for {sleep_time:.2f} seconds...")
        time.sleep(sleep_time)

    print("Max retries reached. Mission failed.")
    return None


def get_with_backoff(driver: WebDriver, url: str, max_retries=5, initial_delay=2):
    retries = 0
    delay = initial_delay

    while retries < max_retries:
        try:
            driver.get(url)
            return True  # Exit function on success

        except TimeoutException:
            retries += 1
            if retries == max_retries:
                print("Max retries reached. Failing...")
                raise

            print(f"Timeout! Retrying [{retries+1}] in {delay} seconds...")
            time.sleep(delay)
            delay *= 2  # Exponentially increase the wait (2, 4, 8, 16...)


def save_xlsx(xlsx_path: str, funds: list[dict], cols: list[str], sheet: str, start: int = 2):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet]
    for fund in funds:
        for idx, val in enumerate(cols):
            col = idx+1
            row = fund.get("index")
            if row:
                start = row
            if val == "url":
                cell = ws.cell(start, col, fund.get(val))
                cell.style = "Hyperlink"
                cell.hyperlink = fund.get(val)
                continue
            ws.cell(start, col, fund.get(val))
        start += 1
    wb.save(xlsx_path)
    wb.close()
